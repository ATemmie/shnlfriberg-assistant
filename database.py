from pathlib import Path
from openpyxl import load_workbook


BOOL_TRUE = {"yes", "true", "1", "active", "是"}


class Player:
    __slots__ = (
        "id", "name", "country", "region", "team", "age",
        "role", "major_wins", "major_apps", "is_active",
    )

    def __init__(self, values: list):
        self.id = int(values[0]) if values[0] is not None else 0
        self.name = str(values[1]) if values[1] else ""
        self.country = str(values[2]) if values[2] else ""
        self.region = str(values[3]) if values[3] else ""
        self.team = str(values[4]) if values[4] else ""
        self.age = int(values[5]) if values[5] is not None else 0
        self.role = str(values[6]) if values[6] else ""
        self.major_wins = int(values[7]) if values[7] is not None else -1
        self.major_apps = int(values[8]) if values[8] is not None else -1
        raw_active = values[9]
        if isinstance(raw_active, bool):
            self.is_active = raw_active
        elif isinstance(raw_active, str):
            self.is_active = raw_active.strip().lower() in BOOL_TRUE
        else:
            self.is_active = False

    def __repr__(self):
        return f"<Player #{self.id}: {self.name}>"


class Database:
    def __init__(self, xlsx_path: str):
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Empty spreadsheet")

        self._players = []
        for vals in rows[1:]:
            p = Player(list(vals))
            if p.name:
                self._players.append(p)

    @property
    def players(self):
        return self._players

    def __len__(self):
        return len(self._players)

    def by_name(self, name: str):
        name_lower = name.strip().lower()
        for p in self._players:
            if p.name.lower() == name_lower:
                return p
        return None
